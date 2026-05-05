"""
build_issues_xlsx.py
====================
Convert REPO_ISSUES_REPORT.md into a color-coded, filterable Excel workbook
for easy issue tracking.

Output workbook (REPO_ISSUES_REPORT.xlsx) contains:
  * "All Issues"  — every issue across every section, with section context
  * "By Section"  — one sheet per markdown section (e.g. "1. Security")
  * "Roll-ups"    — severity / status / contribution summary tables
  * "Change Log"  — the markdown change-log table

Color codes:
  Pri:     🔴 red, 🟠 orange, 🟡 yellow, 🟢 green
  Status:  Open=red, WIP=blue, Fixed=green, Won't Fix=gray

Usage:
  python scripts/build_issues_xlsx.py
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "REPO_ISSUES_REPORT.md"
OUT = ROOT / "REPO_ISSUES_REPORT.xlsx"

# ── Color palette ─────────────────────────────────────────────────────────────
PRI_FILL = {
    "🔴": PatternFill("solid", fgColor="F8D7DA"),   # Critical — red
    "🟠": PatternFill("solid", fgColor="FFE0B2"),   # High     — orange
    "🟡": PatternFill("solid", fgColor="FFF3CD"),   # Medium   — yellow
    "🟢": PatternFill("solid", fgColor="D4EDDA"),   # Low      — green
}
PRI_LABEL = {
    "🔴": "Critical",
    "🟠": "High",
    "🟡": "Medium",
    "🟢": "Low",
}
STATUS_FILL = {
    "Open":      PatternFill("solid", fgColor="F8D7DA"),
    "WIP":       PatternFill("solid", fgColor="CCE5FF"),
    "Fixed":     PatternFill("solid", fgColor="D4EDDA"),
    "Won't Fix": PatternFill("solid", fgColor="E2E3E5"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, color="1F4E78", size=13)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Markdown parsing ─────────────────────────────────────────────────────────
ISSUE_HDR_RE = re.compile(r"^\|\s*#\s*\|\s*Pri\s*\|", re.IGNORECASE)
SECTION_RE = re.compile(r"^##\s+(.+)$")
LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")


def strip_md(cell: str) -> str:
    """Convert markdown link syntax + inline code to plain text."""
    s = cell.strip()
    s = LINK_RE.sub(r"\1", s)
    s = s.replace("`", "")
    return s


def normalize_status(raw: str) -> str:
    raw = raw.strip()
    if "Won't Fix" in raw or "WontFix" in raw or "wont fix" in raw.lower():
        return "Won't Fix"
    if "Fixed" in raw or "✅" in raw:
        return "Fixed"
    if "WIP" in raw or "🔧" in raw or "Progress" in raw:
        return "WIP"
    if "Open" in raw or "⏳" in raw:
        return "Open"
    return raw


def normalize_pri(raw: str) -> str:
    """Pluck the first colored circle. Some rows use a literal Unicode
    'large circle' character that survived earlier edits; map it too."""
    s = raw.strip()
    for ch in ("🔴", "🟠", "🟡", "🟢"):
        if ch in s:
            return ch
    # Heuristic fallback for unknown glyphs — bucket as Medium.
    return "🟡"


def parse_table_row(line: str) -> List[str]:
    parts = [p.strip() for p in line.strip().strip("|").split("|")]
    return parts


def parse_markdown(md: str) -> Tuple[List[dict], List[dict], List[List[str]]]:
    """Return (issues, rollups, change_log)."""
    issues: List[dict] = []
    change_log: List[List[str]] = []
    rollups: List[dict] = []

    lines = md.splitlines()
    i = 0
    current_section = ""
    in_change_log = False
    while i < len(lines):
        line = lines[i]
        m_section = SECTION_RE.match(line)
        if m_section:
            current_section = m_section.group(1).strip()
            in_change_log = current_section.lower().startswith("change log")
            i += 1
            continue

        # Change Log table — the very first table after the "Change Log" heading
        if in_change_log and line.startswith("| Date "):
            # skip header + separator
            i += 2
            while i < len(lines) and lines[i].startswith("|"):
                row = parse_table_row(lines[i])
                if len(row) >= 5:
                    change_log.append([strip_md(c) for c in row[:5]])
                i += 1
            in_change_log = False
            continue

        # Issue tables — header row contains "# | Pri | File | Finding | Status | Owner | Requirements | Analyst Note"
        if ISSUE_HDR_RE.match(line):
            header = parse_table_row(line)
            i += 2  # skip header + separator
            while i < len(lines) and lines[i].startswith("|"):
                row = parse_table_row(lines[i])
                if len(row) >= 8:
                    issues.append({
                        "Section":      current_section,
                        "ID":           strip_md(row[0]),
                        "Pri":          normalize_pri(row[1]),
                        "Severity":     PRI_LABEL.get(normalize_pri(row[1]), ""),
                        "File":         strip_md(row[2]),
                        "Finding":      strip_md(row[3]),
                        "Status":       normalize_status(row[4]),
                        "Owner":        strip_md(row[5]),
                        "Requirements": strip_md(row[6]),
                        "Analyst Note": strip_md(row[7]),
                    })
                i += 1
            continue
        i += 1
    return issues, rollups, change_log


# ── Workbook builders ────────────────────────────────────────────────────────
ISSUE_COLS = [
    ("ID", 7),
    ("Pri", 8),
    ("Severity", 11),
    ("Section", 32),
    ("File", 50),
    ("Finding", 70),
    ("Status", 12),
    ("Owner", 11),
    ("Requirements", 60),
    ("Analyst Note", 90),
]


def write_issue_header(ws, row=1):
    for col_idx, (name, width) in enumerate(ISSUE_COLS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_issue_row(ws, row_idx: int, issue: dict):
    values = [issue.get(name, "") for name, _ in ISSUE_COLS]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = WRAP if col_idx in (4, 5, 6, 9, 10) else CENTER
        cell.border = BORDER
    pri_fill = PRI_FILL.get(issue["Pri"])
    if pri_fill:
        ws.cell(row=row_idx, column=2).fill = pri_fill
        ws.cell(row=row_idx, column=3).fill = pri_fill
    status_fill = STATUS_FILL.get(issue["Status"])
    if status_fill:
        ws.cell(row=row_idx, column=7).fill = status_fill


def add_table(ws, last_row: int, last_col: int, name: str, start_row: int = 1):
    """Wrap the data in an Excel Table so the user gets filter dropdowns."""
    if last_row <= start_row:
        return
    ref = f"A{start_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_all_issues_sheet(wb: Workbook, issues: List[dict]):
    ws = wb.create_sheet("All Issues")
    write_issue_header(ws)
    for idx, issue in enumerate(issues, start=2):
        write_issue_row(ws, idx, issue)
    add_table(ws, last_row=len(issues) + 1, last_col=len(ISSUE_COLS), name="AllIssues")


def build_section_sheets(wb: Workbook, issues: List[dict]):
    by_section: dict[str, List[dict]] = {}
    for it in issues:
        by_section.setdefault(it["Section"], []).append(it)
    for section_idx, (section, items) in enumerate(by_section.items(), start=1):
        # Worksheet names: max 31 chars, no [ ] : * ? / \
        title = re.sub(r"[\[\]:*?/\\]", "", section)[:31] or f"Section {section_idx}"
        ws = wb.create_sheet(title)
        ws.cell(row=1, column=1, value=section).font = SECTION_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ISSUE_COLS))
        write_issue_header(ws, row=2)
        for idx, issue in enumerate(items, start=3):
            write_issue_row(ws, idx, issue)
        add_table(
            ws,
            last_row=len(items) + 2,
            last_col=len(ISSUE_COLS),
            name=f"Sec{section_idx}",
            start_row=2,
        )
        # Re-freeze below the header
        ws.freeze_panes = ws.cell(row=3, column=1)


def build_rollups_sheet(wb: Workbook, issues: List[dict]):
    ws = wb.create_sheet("Roll-ups")
    # Severity roll-up
    sev_order = ["🔴", "🟠", "🟡", "🟢"]
    sev_counts = {p: {"Resolved": 0, "Won't Fix": 0, "Open": 0, "WIP": 0} for p in sev_order}
    for it in issues:
        bucket = sev_counts.setdefault(
            it["Pri"], {"Resolved": 0, "Won't Fix": 0, "Open": 0, "WIP": 0}
        )
        if it["Status"] == "Fixed":
            bucket["Resolved"] += 1
        elif it["Status"] == "Won't Fix":
            bucket["Won't Fix"] += 1
        elif it["Status"] == "WIP":
            bucket["WIP"] += 1
        else:
            bucket["Open"] += 1

    ws.cell(row=1, column=1, value="Severity Roll-up").font = SECTION_FONT
    headers = ["Pri", "Severity", "Total", "Resolved", "Won't Fix", "WIP", "Open"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    row = 3
    grand = {"Total": 0, "Resolved": 0, "Won't Fix": 0, "WIP": 0, "Open": 0}
    for pri in sev_order:
        c = sev_counts[pri]
        total = sum(c.values())
        ws.cell(row=row, column=1, value=pri).alignment = CENTER
        ws.cell(row=row, column=2, value=PRI_LABEL[pri]).alignment = CENTER
        for col, key in enumerate(["Total", "Resolved", "Won't Fix", "WIP", "Open"], start=3):
            v = total if key == "Total" else c[key]
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = CENTER
            cell.border = BORDER
        for col in (1, 2):
            ws.cell(row=row, column=col).fill = PRI_FILL[pri]
            ws.cell(row=row, column=col).border = BORDER
        for k in grand:
            grand[k] += total if k == "Total" else c[k]
        row += 1
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
    for col, key in enumerate(["Total", "Resolved", "Won't Fix", "WIP", "Open"], start=3):
        cell = ws.cell(row=row, column=col, value=grand[key])
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = BORDER

    # Status roll-up
    status_counts = {"Open": 0, "WIP": 0, "Fixed": 0, "Won't Fix": 0}
    for it in issues:
        status_counts[it["Status"]] = status_counts.get(it["Status"], 0) + 1
    start = row + 3
    ws.cell(row=start, column=1, value="Status Roll-up").font = SECTION_FONT
    for c, h in enumerate(["Status", "Count"], start=1):
        cell = ws.cell(row=start + 1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for r, (status, count) in enumerate(status_counts.items(), start=start + 2):
        cell_status = ws.cell(row=r, column=1, value=status)
        cell_status.alignment = CENTER
        cell_status.border = BORDER
        if STATUS_FILL.get(status):
            cell_status.fill = STATUS_FILL[status]
        cell_count = ws.cell(row=r, column=2, value=count)
        cell_count.alignment = CENTER
        cell_count.border = BORDER

    # Contribution tracker
    owners: dict[str, dict] = {}
    for it in issues:
        owner = it["Owner"] or "—"
        if owner in ("—", ""):
            owner = "— (unassigned)"
        bucket = owners.setdefault(owner, {"Open": 0, "WIP": 0, "Fixed": 0, "Won't Fix": 0})
        bucket[it["Status"]] = bucket.get(it["Status"], 0) + 1
    start = r + 3
    ws.cell(row=start, column=1, value="Contribution Tracker").font = SECTION_FONT
    for c, h in enumerate(["Owner", "Open", "WIP", "Fixed", "Won't Fix", "Total Touched"], start=1):
        cell = ws.cell(row=start + 1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    r2 = start + 2
    for owner, b in sorted(owners.items()):
        ws.cell(row=r2, column=1, value=owner).alignment = CENTER
        for col, key in enumerate(["Open", "WIP", "Fixed", "Won't Fix"], start=2):
            cell = ws.cell(row=r2, column=col, value=b.get(key, 0))
            cell.alignment = CENTER
            cell.border = BORDER
        ws.cell(row=r2, column=6, value=sum(b.values())).alignment = CENTER
        for c in range(1, 7):
            ws.cell(row=r2, column=c).border = BORDER
        r2 += 1

    for col_idx, width in enumerate([22, 14, 10, 12, 14, 8, 14], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def build_change_log_sheet(wb: Workbook, change_log: List[List[str]]):
    if not change_log:
        return
    ws = wb.create_sheet("Change Log")
    headers = ["Date", "Change", "Issues Affected", "Files Touched", "Author"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for r, row in enumerate(change_log, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
    for col_idx, width in enumerate([14, 80, 28, 50, 12], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    add_table(ws, last_row=len(change_log) + 1, last_col=5, name="ChangeLog")


def build_legend_sheet(wb: Workbook):
    ws = wb.create_sheet("Legend", 0)  # first sheet
    ws.cell(row=1, column=1, value="JobPilot — Issue Tracker").font = Font(bold=True, size=16, color="1F4E78")
    ws.cell(row=2, column=1, value=f"Generated from REPO_ISSUES_REPORT.md").font = Font(italic=True, color="666666")

    ws.cell(row=4, column=1, value="Priority Color Codes").font = SECTION_FONT
    pri_rows = [
        ("🔴", "Critical", "Production-blocking; fix this sprint."),
        ("🟠", "High",     "Significant defect; fix next sprint."),
        ("🟡", "Medium",   "Quality / maintainability; planned work."),
        ("🟢", "Low",      "Nice-to-have / cleanup."),
    ]
    for r, (icon, label, desc) in enumerate(pri_rows, start=5):
        ws.cell(row=r, column=1, value=icon).alignment = CENTER
        ws.cell(row=r, column=1).fill = PRI_FILL[icon]
        ws.cell(row=r, column=2, value=label).font = Font(bold=True)
        ws.cell(row=r, column=3, value=desc).alignment = WRAP

    ws.cell(row=10, column=1, value="Status Color Codes").font = SECTION_FONT
    status_rows = [
        ("Open",      "Not yet started."),
        ("WIP",       "In progress (limit one per owner where possible)."),
        ("Fixed",     "Resolved and verified."),
        ("Won't Fix", "Acknowledged; deferred per documented requirements."),
    ]
    for r, (label, desc) in enumerate(status_rows, start=11):
        cell = ws.cell(row=r, column=1, value=label)
        cell.alignment = CENTER
        cell.font = Font(bold=True)
        cell.fill = STATUS_FILL[label]
        ws.cell(row=r, column=3, value=desc).alignment = WRAP

    ws.cell(row=16, column=1, value="Sheets in this workbook").font = SECTION_FONT
    sheet_rows = [
        ("All Issues",   "Master table — every issue, filterable."),
        ("By Section",   "One sheet per markdown section (Security, API Design, Frontend, etc.)."),
        ("Roll-ups",     "Severity / status / contribution summaries."),
        ("Change Log",   "Sprint-by-sprint history from the markdown report."),
    ]
    for r, (name, desc) in enumerate(sheet_rows, start=17):
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        ws.cell(row=r, column=3, value=desc).alignment = WRAP

    for col_idx, width in enumerate([14, 16, 80], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main() -> int:
    md = SRC.read_text(encoding="utf-8")
    issues, _rollups, change_log = parse_markdown(md)

    wb = Workbook()
    # Drop the default sheet — we add Legend first, then the rest.
    wb.remove(wb.active)

    build_legend_sheet(wb)
    build_all_issues_sheet(wb, issues)
    build_section_sheets(wb, issues)
    build_rollups_sheet(wb, issues)
    build_change_log_sheet(wb, change_log)

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(issues)} issues across {len({i['Section'] for i in issues})} sections)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
